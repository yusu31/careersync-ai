"""
追加情報（テキスト・URL・ファイル）を元に企業情報を補完するモジュール。

対応入力:
  - テキスト（面接メモ・コピペ情報など自由入力）
  - URL（複数）: scraper でテキスト抽出後に送信
  - 画像 (.jpg/.png/.gif/.webp): Gemini Vision にバイナリ直接送信
  - PDF (.pdf): Gemini PDF ネイティブサポートで直接送信
  - Word (.docx): python-docx でテキスト抽出
  - Excel (.xlsx/.xls): openpyxl でシート内容をテキスト化
"""

import json
import io
from dataclasses import dataclass
from typing import Optional

import google.generativeai as genai
from config.settings import settings
from services.scraper import scrape_company

genai.configure(api_key=settings.gemini_api_key)

_MODEL = genai.GenerativeModel(
    model_name="gemini-2.5-flash",
    generation_config=genai.types.GenerationConfig(
        response_mime_type="application/json",
        temperature=0.2,
    ),
)

# Gemini に直接送れるバイナリ形式
_BINARY_MIME_TYPES = {
    "image/jpeg", "image/png", "image/gif", "image/webp",
    "application/pdf",
}

# ファイル拡張子 → MIME タイプのマッピング
_EXT_TO_MIME = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
}


@dataclass
class UploadedFile:
    filename: str
    content: bytes
    mime_type: str


def _extract_docx_text(content: bytes) -> str:
    """Word (.docx) からテキストを抽出する。"""
    try:
        from docx import Document
        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        return f"[Word ファイルの読み込みに失敗しました: {e}]"


def _extract_excel_text(content: bytes) -> str:
    """Excel (.xlsx/.xls) からテキストを抽出する。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"[シート: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                row_text = "  |  ".join(str(v) for v in row if v is not None)
                if row_text.strip():
                    lines.append(row_text)
        return "\n".join(lines)
    except Exception as e:
        return f"[Excel ファイルの読み込みに失敗しました: {e}]"


def _build_parts(
    text: str,
    urls: list[str],
    files: list[UploadedFile],
    current_info: dict,
) -> list:
    """Gemini に送信するコンテンツパーツのリストを構築する。"""
    company_json = json.dumps(
        {k: v for k, v in current_info.items() if v is not None},
        ensure_ascii=False, indent=2
    )

    system_prompt = f"""あなたは転職活動支援AIです。

現在登録されている企業「{current_info.get('name', '不明')}」の情報:
{company_json}

以下の追加情報を元に、企業情報を更新・補完してください。
確実に正しいと判断できるフィールドのみ返してください。
更新不要なフィールドは省略してください。

出力は下記キーを使った JSON のみ返してください（コードブロック不要）:
name, industry, employees, founded_year, listing_status, development_type,
location, work_style, overtime_hours, paid_leave_rate, transfer,
salary, expected_first_salary, salary_upper, inexperienced_ok, training_program,
hiring_probability_score, job_description, skill_stack, tech_growth_score,
career_growth_score, career_path, benefits, summary,
strengths_weaknesses（{{strengths:[...], weaknesses:[...]}}）,
interview_strategy（[{{question, answer}},...] の配列）,
scores（{{growth, stability, culture_fit, work_life_balance, compensation}} 各1〜10）

出力例:
{{
  "summary": "...",
  "interview_strategy": [{{"question": "...", "answer": "..."}}]
}}"""

    parts: list = [system_prompt]

    if text and text.strip():
        parts.append(f"\n[ユーザーテキスト情報]\n{text.strip()}")

    for url in urls:
        if url.strip():
            scraped = scrape_company(url.strip())
            parts.append(f"\n[URL: {url}]\n{scraped or '（スクレイピング失敗）'}")

    binary_parts = []
    for f in files:
        if f.mime_type in _BINARY_MIME_TYPES:
            binary_parts.append({"mime_type": f.mime_type, "data": f.content})
        elif "wordprocessingml" in f.mime_type:
            extracted = _extract_docx_text(f.content)
            parts.append(f"\n[ドキュメント: {f.filename}]\n{extracted}")
        elif "spreadsheet" in f.mime_type or "ms-excel" in f.mime_type:
            extracted = _extract_excel_text(f.content)
            parts.append(f"\n[スプレッドシート: {f.filename}]\n{extracted}")

    # バイナリ（画像・PDF）は末尾に追加（Gemini の推奨順序）
    parts.extend(binary_parts)
    return parts


def supplement(
    current_info: dict,
    text: str = "",
    urls: Optional[list[str]] = None,
    files: Optional[list[UploadedFile]] = None,
) -> dict:
    """
    追加情報を Gemini に送り、更新すべきフィールドと更新後の企業情報を返す。

    返却:
        {
            "updated_fields": ["summary", "scores", ...],
            "updates": { フィールド名: 値, ... }   # DB更新用
        }
    """
    urls = urls or []
    files = files or []

    parts = _build_parts(text, urls, files, current_info)

    response = _MODEL.generate_content(parts)
    raw = response.text.strip()

    # ```json ... ``` で返ってきた場合に備えて除去
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]

    ai_result: dict = json.loads(raw.strip())

    # オブジェクト型フィールドは JSON 文字列に変換（DB 保存形式）
    _JSON_FIELDS = {"strengths_weaknesses", "interview_strategy", "scores", "skill_stack"}
    updates = {}
    for k, v in ai_result.items():
        if v is None:
            continue
        if k in _JSON_FIELDS and isinstance(v, (dict, list)):
            updates[k] = json.dumps(v, ensure_ascii=False)
        else:
            updates[k] = v

    return {
        "updated_fields": list(updates.keys()),
        "updates": updates,
    }
